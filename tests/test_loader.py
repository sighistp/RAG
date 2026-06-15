import pytest
from rag.loader import load


def test_load_txt(tmp_path):
    file = tmp_path / "doc.txt"
    file.write_text("hello world", encoding="utf-8")
    assert load(str(file)) == "hello world"


def test_load_markdown(tmp_path):
    file = tmp_path / "doc.md"
    file.write_text("# Title\ncontent", encoding="utf-8")
    result = load(str(file))
    assert "# Title" in result
    assert "content" in result


def test_load_chinese(tmp_path):
    file = tmp_path / "cn.txt"
    file.write_text("你好世界", encoding="utf-8")
    assert load(str(file)) == "你好世界"


def test_load_nonexistent():
    with pytest.raises(FileNotFoundError):
        load("/nonexistent/file.txt")


def test_load_unsupported_format(tmp_path):
    file = tmp_path / "doc.xyz"
    file.write_text("test", encoding="utf-8")
    with pytest.raises(ValueError, match="不支持的文件格式"):
        load(str(file))


def test_load_excel(tmp_path):
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.append(["Name", "Age"])
    ws.append(["Alice", 30])
    ws.append(["Bob", 25])
    file = tmp_path / "data.xlsx"
    wb.save(str(file))
    result = load(str(file))
    assert "Name" in result
    assert "Alice" in result
    assert "30" in result


def test_load_excel_multiple_sheets(tmp_path):
    from openpyxl import Workbook
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Sheet1"
    ws1.append(["A", "B"])
    ws2 = wb.create_sheet("Sheet2")
    ws2.append(["X", "Y"])
    file = tmp_path / "multi.xlsx"
    wb.save(str(file))
    result = load(str(file))
    assert "Sheet: Sheet1" in result
    assert "Sheet: Sheet2" in result
    assert "A" in result
    assert "X" in result


def test_load_excel_empty(tmp_path):
    from openpyxl import Workbook
    wb = Workbook()
    file = tmp_path / "empty.xlsx"
    wb.save(str(file))
    result = load(str(file))
    assert "=== Sheet:" in result


def test_load_excel_corrupt_file(tmp_path):
    file = tmp_path / "bad.xlsx"
    file.write_bytes(b"not a valid excel file")
    with pytest.raises(ValueError, match="无法读取 Excel 文件"):
        load(str(file))
