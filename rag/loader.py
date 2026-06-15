import os
import shutil
import tempfile
import warnings
from pathlib import Path

from rag.cleaner import clean_text, detect_and_decode

SUPPORTED_EXTENSIONS = {".txt", ".md", ".pdf", ".docx", ".xlsx"}


def load(file_path: str) -> str:
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"文件不存在: {file_path}")
    ext = os.path.splitext(file_path)[1].lower()
    if ext not in SUPPORTED_EXTENSIONS:
        raise ValueError(f"不支持的文件格式: {ext}")
    if ext == ".pdf":
        return load_pdf(file_path)
    if ext == ".docx":
        return load_docx(file_path)
    if ext == ".xlsx":
        return load_excel(file_path)
    # .txt, .md files: detect encoding + clean
    with open(file_path, "rb") as f:
        raw = f.read()
    text = detect_and_decode(raw)
    return clean_text(text)


def _ensure_java_on_path():
    """Put Java on PATH if not already accessible (for OpenDataLoader PDF)."""
    import shutil

    if shutil.which("java"):
        return
    candidates = [
        os.environ.get("JAVA_HOME", ""),
        r"C:\Program Files\Amazon Corretto\jdk17.0.19_10",
        r"C:\Program Files\Java\jdk-17",
        r"C:\Program Files\Eclipse Adoptium\jdk-17",
    ]
    for base in candidates:
        if not base:
            continue
        java_bin = os.path.join(base, "bin")
        if os.path.isfile(os.path.join(java_bin, "java.exe")):
            os.environ["PATH"] = java_bin + os.pathsep + os.environ["PATH"]
            return


def load_pdf(file_path: str) -> str:
    """Extract text from PDF using OpenDataLoader (converts to Markdown)."""
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"文件不存在: {file_path}")

    import opendataloader_pdf

    _ensure_java_on_path()
    tmp_dir = tempfile.mkdtemp()
    try:
        opendataloader_pdf.convert(
            input_path=[file_path],
            output_dir=tmp_dir,
            format="markdown",
        )
        stem = Path(file_path).stem
        md_path = Path(tmp_dir) / f"{stem}.md"
        if md_path.exists():
            return md_path.read_text(encoding="utf-8")
        warnings.warn(f"OpenDataLoader PDF 未生成输出文件: {md_path}")
        return ""
    finally:
        shutil.rmtree(tmp_dir, ignore_errors=True)


def load_docx(file_path: str) -> str:
    """Extract text from DOCX using python-docx (paragraphs + tables in order)."""
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"文件不存在: {file_path}")
    from docx import Document
    from docx.oxml.ns import qn

    doc = Document(file_path)
    parts = []
    for elem in doc.element.body:
        if elem.tag == qn("w:p"):
            texts = [t.text for t in elem.iter(qn("w:t")) if t.text]
            line = "".join(texts).strip()
            if line:
                parts.append(line)
        elif elem.tag == qn("w:tbl"):
            for row in elem.iter(qn("w:tr")):
                cells = []
                for cell in row.iter(qn("w:tc")):
                    texts = [t.text for t in cell.iter(qn("w:t")) if t.text]
                    cells.append("".join(texts).strip())
                line = " | ".join(cells)
                if line.strip():
                    parts.append(line)
    return "\n".join(parts)


def load_excel(file_path: str) -> str:
    """Read .xlsx file, convert all sheets to text."""
    from openpyxl import load_workbook

    try:
        wb = load_workbook(file_path, read_only=True, data_only=True)
    except Exception as e:
        raise ValueError(f"无法读取 Excel 文件: {file_path} ({e})")
    parts = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        parts.append(f"=== Sheet: {sheet} ===")
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) if c is not None else "" for c in row]
            parts.append("\t".join(cells))
    wb.close()
    return "\n".join(parts)
